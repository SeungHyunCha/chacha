import openpyxl

class ParsingExcel:
    def __init__(self, name, change):
        self.name = name
        self.change = change
        self.book = None
        self.__getWorkbook()
        
    def __getWorkbook(self):
        self.book = openpyxl.load_workbook(self.name)
        
    def getS1(self):
        return self.book['Sheet1']
    
    def getS2(self):
        return self.book['Sheet2']
    
    def getS3(self):
        return self.book['Sheet3']
    
    def getS4(self):
        return self.book['Sheet4']
    
    def getS5(self):
        return self.book['Sheet5']
    
    def getDeprecation(self, sheet):
        for mycell in sheet.rows:
            for col_num in range(sheet.max_column):
                if mycell[col_num].value == 'deprecation':
                    dep_list = {}
                    # name
                    dep_name = mycell[col_num].value
                    dep_list.update({1:dep_name})
                    # column
                    dep_al = chr(65+col_num)
                    dep_list.update({0:dep_al})
                    # row
                    dep_row = len(sheet[dep_al])
                    for i in range(2,dep_row+1):
                        dep_list.update({i:sheet[dep_al + str(i)].value})
        
        return dep_list
    
    def getConvert(self, sheet):
        for mycell in sheet.rows:
            for col_num in range(sheet.max_column):
                if mycell[col_num].value == 'convert':
                    con_list = {}
                    #name
                    con_name = mycell[col_num].value
                    con_list.update({1:con_name})
                    #column
                    con_al = chr(65+col_num)
                    con_list.update({0:con_al})
                    #row
                    con_row= len(sheet[con_al])
                    for i in range(2,con_row+1):
                        con_list.update({i:sheet[con_al + str(i)].value})
        
        return con_list
    
    def reasonforChange(self, sheet, dep, con):
        for key, value in dep.items():
            if value == 'D' or value == 'N': 
                con[key] = self.change
                sheet[con[0] + str(key)].value = self.change 
            else: pass
        self.book.save(self.name)
        return con

if __name__ == '__main__':
    excel = 'test.xlsx'
    parser = ParsingExcel(excel, 'test')
    s1 = parser.getS1()
    dep = parser.getDeprecation(s1)
    print dep
    con = parser.getConvert(s1)
    print con
    comp = parser.reasonforChange(s1, dep, con)
    print comp
    